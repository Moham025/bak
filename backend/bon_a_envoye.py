from flask import Flask, request, send_file, jsonify
# from numpy import array_str # Removing this conflicting import
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
import os
import re
import io
from copy import copy # Pour copier les styles
import math # Pour math.trunc
from flask_cors import CORS # S'assurer que l'import est présent
import sys

# Ajouter le répertoire backend au path pour les imports
backend_dir = os.path.dirname(os.path.abspath(__file__))
if backend_dir not in sys.path:
    sys.path.insert(0, backend_dir)

# Import des modules du répertoire backend
from covnumletter import conv_number_letter as cl_conv_number_letter
from combineArm import process_armature_csvs

# Import des modules EstimBatiment
estim_batiment_dir = os.path.join(backend_dir, 'EstimBatiment')
if estim_batiment_dir not in sys.path:
    sys.path.insert(0, estim_batiment_dir)

from data_reader import get_qt_data, get_open_data, get_simple_block_data, get_formula_block_data
from calculation_engine import parse_calcul_sheet_and_process_blocks, process_menuiserie_block, process_simple_block, process_formula_block, write_recap_block

# --- Flask App Setup ---
app = Flask(__name__)
CORS(app) # Décommentez si vous avez des appels cross-origin et que vous voulez les gérer

# --- Fonctions Utilitaires Excel Python ---
def trouver_nom_feuille_original(nom_saisi, noms_feuilles_sources_dict):
    nom_normalise_saisi = nom_saisi.strip().lower()
    return noms_feuilles_sources_dict.get(nom_normalise_saisi, "")

def feuille_existe(workbook, sheet_name):
    return sheet_name in workbook.sheetnames

def est_une_feuille_recap(nom_feuille_original):
    nom_lower = nom_feuille_original.strip().lower()
    recap_keywords = ["recap", "récap", "summary", "synthese", "synthèse"]
    specific_recap_names = ["recapitulatif", "récapitulatif"]
    is_recap = any(keyword in nom_lower for keyword in recap_keywords) or nom_lower in specific_recap_names
    print(f"  Vérif Recap pour '{nom_feuille_original}' (normalisé: '{nom_lower}'): {is_recap}")
    return is_recap

def add_quotes_if_necessary(sheet_name):
    if re.search(r"[\s!@#$%^&*()+={}\[\]:;\"'<>,.?/\\|-]", sheet_name) or \
       sheet_name.lower() in ['true', 'false'] or \
       (sheet_name and sheet_name[0].isdigit()):
        return f"'{sheet_name.replace('\'', '\'\'')}'"
    return sheet_name

# --- Fonctions de Traitement de Feuilles Excel Python ---
def nettoyer_total_en_lettres(ws_copie, ws_original_source_values):
    """Similaire à la fonction VBA NettoyerTotalEnLettres."""
    print(f"  Nettoyage du total en lettres pour la feuille '{ws_copie.title}'...")
    prefixe = "Arrêter le présent devis estimatif à la somme de :"
    
    if ws_copie.max_row == 0:
        print("    DEBUG: Feuille vide, pas de nettoyage de total.")
        return

    ligne_total_general = None
    valeur_cell_A_total_general = None
    for row in range(1, ws_copie.max_row + 1):
        cell_a = ws_copie[f"A{row}"]
        if cell_a.value and isinstance(cell_a.value, str):
            if "TOTAL GENERAL" in cell_a.value.upper():
                ligne_total_general = row
                valeur_cell_A_total_general = cell_a.value
                print(f"    DEBUG: Ligne 'TOTAL GENERAL' TROUVÉE. Row: {row}, Contenu: '{valeur_cell_A_total_general}'")
                break
    
    if ligne_total_general:
        cell_f_total_coord = f"F{ligne_total_general}"
        cell_f_total_formule_mode = ws_copie[cell_f_total_coord] # From the sheet with formulas
        cell_f_total_valeur_mode = ws_original_source_values[cell_f_total_coord] # From the sheet with values

        print(f"    DEBUG: Cellule du total (col F): {cell_f_total_coord}")
        print(f"      Valeur (depuis feuille formules ws_copie): '{cell_f_total_formule_mode.value}', Type: {cell_f_total_formule_mode.data_type}")
        print(f"      Valeur (depuis feuille valeurs ws_original_source_values): '{cell_f_total_valeur_mode.value}', Type: {cell_f_total_valeur_mode.data_type}")

        montant_total = None
        # Prioriser la valeur de la feuille chargée avec data_only=True pour les formules
        if cell_f_total_formule_mode.data_type == 'f':
            print(f"    DEBUG: {cell_f_total_coord} est une formule ('{cell_f_total_formule_mode.value}'). Utilisation de la valeur de ws_original_source_values.")
            montant_total = cell_f_total_valeur_mode.value
        else:
            # Si ce n'est pas une formule, la valeur de ws_copie devrait être correcte
            montant_total = cell_f_total_formule_mode.value
            print(f"    DEBUG: {cell_f_total_coord} n'est pas une formule. Utilisation de la valeur de ws_copie: {montant_total}")

        if isinstance(montant_total, (int, float)) and montant_total > 0:
            print(f"    DEBUG: Montant total final utilisé pour conversion: {montant_total}")
            
            texte_total_lettres = cl_conv_number_letter(montant_total, devise=1, langue=0)
            texte_final = f"{prefixe} {texte_total_lettres}"
            ligne_suivante = ligne_total_general + 1
            cell_a_suivante = ws_copie[f"A{ligne_suivante}"]
            cell_a_suivante.value = texte_final
            
            print(f"    DEBUG: Texte en lettres ('{texte_total_lettres}') inséré ligne {ligne_suivante} col A avec préfixe.")
        elif montant_total is None and cell_f_total_formule_mode.data_type == 'f':
            print(f"    AVERTISSEMENT: Montant total dans {cell_f_total_coord} (formule: '{cell_f_total_formule_mode.value}') n'a pas pu être résolu en nombre depuis la feuille des valeurs ('{cell_f_total_valeur_mode.value}'). Conversion ignorée.")
        else:
            print(f"    AVERTISSEMENT: Valeur montant invalide ou nulle dans {cell_f_total_coord}: '{montant_total}'. Conversion ignorée.")
    else:
        print("    DEBUG: Aucune ligne 'TOTAL GENERAL' trouvée dans la colonne A.")

    print(f"  Fin nettoyage total en lettres pour '{ws_copie.title}'.")

def modifier_liens_externes_feuille_recap(ws_recap, wb_cible):
    print(f"--- Début ModifierLiensExternesFeuilleRecap pour '{ws_recap.title}' ---")
    
    # Analyser plusieurs colonnes : F et la dernière colonne
    colonnes_a_analyser = []
    
    # Toujours analyser la colonne F
    colonnes_a_analyser.append(6)  # F = colonne 6
    
    # Analyser aussi la dernière colonne si elle existe et est différente de F
    if ws_recap.max_column > 0 and ws_recap.max_column != 6:
        colonnes_a_analyser.append(ws_recap.max_column)
    
    if not colonnes_a_analyser:
        print(f"    Aucune colonne à analyser pour '{ws_recap.title}'. Arrêt.")
        print(f"--- Fin ModifierLiensExternesFeuilleRecap pour '{ws_recap.title}' ---")
        return
    
    for col_num in colonnes_a_analyser:
        col_letter = get_column_letter(col_num)
        print(f"    Analyse des formules dans la colonne {col_letter} de '{ws_recap.title}'")

        for r in range(1, ws_recap.max_row + 1):
            cell = ws_recap.cell(row=r, column=col_num)
            if cell.data_type == 'f':
                formula_string = str(cell.value)
                print(f"      Analyse cellule {cell.coordinate} | Formule Originale: {formula_string}")
                
                # Traiter les liens externes complexes avec fichier externe [nom_fichier]
                match_externe = re.match(r"=(.*?)\[([^\]]+)\](.*?([^\!']+?)|\'?([^\!']+?)\'?)\!(.+)", formula_string, re.IGNORECASE)
                if match_externe:
                    prefix, external_file, _, _, external_sheet_raw, cell_ref = match_externe.groups()
                    external_sheet_clean = external_sheet_raw.strip("'").replace("''", "'").strip()
                    print(f"        Lien externe: Fichier='{external_file}', Feuille='{external_sheet_clean}', Cellule='{cell_ref}'")
                    target_copied_sheet_name = f"{external_sheet_clean}_copie"
                    print(f"          Cible attendue: '{target_copied_sheet_name}'")
                    if target_copied_sheet_name in wb_cible.sheetnames:
                        print(f"            Cible '{target_copied_sheet_name}' existe.")
                        new_formula = f"={prefix}{add_quotes_if_necessary(target_copied_sheet_name)}!{cell_ref}"
                        if formula_string.lower() != new_formula.lower():
                            cell.value = new_formula
                            print(f"              MODIFIÉ: '{new_formula}'")
                        else: print("              Aucune modification nécessaire.")
                    else: print(f"            ATTENTION: Cible '{target_copied_sheet_name}' N'EXISTE PAS.")
                else:
                    # Traiter les liens internes simples =NomFeuille!Cellule
                    match_interne = re.match(r"=([^!\[]+)!(.+)", formula_string, re.IGNORECASE)
                    if match_interne:
                        sheet_name_raw, cell_ref = match_interne.groups()
                        sheet_name_clean = sheet_name_raw.strip("'").replace("''", "'").strip()
                        print(f"        Lien interne: Feuille='{sheet_name_clean}', Cellule='{cell_ref}'")
                        target_copied_sheet_name = f"{sheet_name_clean}_copie"
                        print(f"          Cible attendue: '{target_copied_sheet_name}'")
                        if target_copied_sheet_name in wb_cible.sheetnames:
                            print(f"            Cible '{target_copied_sheet_name}' existe.")
                            new_formula = f"={add_quotes_if_necessary(target_copied_sheet_name)}!{cell_ref}"
                            if formula_string.lower() != new_formula.lower():
                                cell.value = new_formula
                                print(f"              MODIFIÉ: '{new_formula}'")
                            else: print("              Aucune modification nécessaire.")
                        else: print(f"            ATTENTION: Cible '{target_copied_sheet_name}' N'EXISTE PAS.")
                    # else: print("        Pas un lien reconnu.") # commenter pour moins de verbosité
    
    print(f"--- Fin ModifierLiensExternesFeuilleRecap pour '{ws_recap.title}' ---")

def copier_feuille_manuellement(ws_source, wb_destination, nouveau_nom_feuille):
    if nouveau_nom_feuille in wb_destination.sheetnames:
        del wb_destination[nouveau_nom_feuille]
    ws_destination = wb_destination.create_sheet(title=nouveau_nom_feuille)
    for row in ws_source.iter_rows():
        for cell in row:
            new_cell = ws_destination.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font, new_cell.border, new_cell.fill = copy(cell.font), copy(cell.border), copy(cell.fill)
                new_cell.number_format, new_cell.protection, new_cell.alignment = cell.number_format, copy(cell.protection), copy(cell.alignment)
    for col_letter, dim in ws_source.column_dimensions.items(): ws_destination.column_dimensions[col_letter].width = dim.width
    for row_num, dim in ws_source.row_dimensions.items(): ws_destination.row_dimensions[row_num].height = dim.height
    for merged_range_str in ws_source.merged_cells.ranges: ws_destination.merge_cells(str(merged_range_str)) # Convertir MergedCellRange en string
    return ws_destination

# --- Logique principale de traitement du serveur ---
def traiter_fichier_excel_core(bytes_fichier_source, noms_feuilles_a_traiter_str):
    try:
        # Ouvrir le fichier deux fois : une pour les formules, une pour les valeurs
        wb_source = openpyxl.load_workbook(io.BytesIO(bytes_fichier_source), data_only=False)
        wb_source_values = openpyxl.load_workbook(io.BytesIO(bytes_fichier_source), data_only=True)
        print(f"Fichier source chargé en mémoire (formules et valeurs).")
    except InvalidFileException:
        print("Fichier Excel invalide ou corrompu.")
        return None
    except Exception as e:
        print(f"Impossible de charger le fichier source depuis les bytes: {e}")
        return None

    noms_feuilles_sources_dict = {name.strip().lower(): name.strip() for name in wb_source.sheetnames}

    if not noms_feuilles_a_traiter_str:
        print("Erreur critique: Noms de feuilles à traiter non fournis.")
        return None # Dans une app web, la sélection de feuilles doit être explicite
        
    l_array_str = [s.strip() for s in noms_feuilles_a_traiter_str.replace(";", ",").split(",") if s.strip()]
    if not l_array_str:
        print("Aucun nom de feuille valide n'a été fourni pour traitement.")
        return None
        
    wb_destination = openpyxl.Workbook()
    default_sheet_name = wb_destination.sheetnames[0] # Généralement "Sheet"
    sheet_to_delete_if_unused = default_sheet_name if default_sheet_name.lower() in ["sheet", "feuil1"] else None
    
    first_sheet_processed = False
    processed_sheet_names_in_dest = []

    for nom_feuille_saisi in l_array_str:
        nom_feuille_source_original = trouver_nom_feuille_original(nom_feuille_saisi, noms_feuilles_sources_dict)
        if nom_feuille_source_original:
            ws_original_source = wb_source[nom_feuille_source_original]
            ws_original_source_values = wb_source_values[nom_feuille_source_original]
            nom_feuille_copie_dest = f"{nom_feuille_source_original.strip()}_copie"
            
            print(f"Copie de la feuille '{ws_original_source.title}' vers '{nom_feuille_copie_dest}'.")
            ws_copie_dest = copier_feuille_manuellement(ws_original_source, wb_destination, nom_feuille_copie_dest)
            
            processed_sheet_names_in_dest.append(nom_feuille_copie_dest)

            if not first_sheet_processed and sheet_to_delete_if_unused and \
               sheet_to_delete_if_unused in wb_destination.sheetnames and \
               nom_feuille_copie_dest != sheet_to_delete_if_unused:
                del wb_destination[sheet_to_delete_if_unused]
                print(f"Feuille par défaut '{sheet_to_delete_if_unused}' supprimée.")
                sheet_to_delete_if_unused = None 
            first_sheet_processed = True
            
            print(f"  Traitement de la feuille copiée: '{ws_copie_dest.title}'")
            
            # 1. Conversion des colonnes D et E en valeurs (AVANT suppression des colonnes G-J)
            print(f"    Conversion des formules en valeurs pour colonnes D et E...")
            for col_letter_idx_str in ["D", "E"]:
                print(f"      Traitement colonne {col_letter_idx_str}...")
                for row in range(1, ws_copie_dest.max_row + 1):
                    cell = ws_copie_dest[f"{col_letter_idx_str}{row}"]
                    if cell.data_type == 'f':  # Si c'est une formule
                        # Récupérer la valeur calculée depuis wb_source_values
                        try:
                            valeur_calculee = ws_original_source_values[f"{col_letter_idx_str}{row}"].value
                            if valeur_calculee is not None:
                                cell.value = valeur_calculee  # Remplacer la formule par la valeur
                                print(f"        {cell.coordinate}: Formule convertie en valeur = {valeur_calculee}")
                            else:
                                print(f"        {cell.coordinate}: Valeur calculée = None, formule conservée")
                        except Exception as e_val:
                            print(f"        Erreur conversion {cell.coordinate}: {e_val}")
                    elif cell.value is not None:
                        # Si ce n'est pas une formule mais a une valeur, la garder
                        print(f"        {cell.coordinate}: Valeur déjà présente = {cell.value}")

            # 1.5. La colonne F garde ses formules intactes
            # Les formules seront recalculées automatiquement par Excel à l'ouverture

            # 2. Nettoyer total en lettres
            nettoyer_total_en_lettres(ws_copie_dest, ws_original_source_values)

            # 3. Supprimer colonnes G à J (APRÈS conversion D et E)
            print(f"    Suppression des colonnes G à J pour {ws_copie_dest.title}...")
            ws_copie_dest.delete_cols(7, 4) # G=7, 4 colonnes (G,H,I,J)

            # 4. Vérifier si c'est un récap et modifier les liens
            nom_original_pour_recap = nom_feuille_source_original.strip()
            if est_une_feuille_recap(nom_original_pour_recap):
                print(f"  Feuille '{ws_copie_dest.title}' identifiée comme récap. Modification des liens...")
                modifier_liens_externes_feuille_recap(ws_copie_dest, wb_destination)
            
            print(f"  Traitement terminé pour '{ws_copie_dest.title}'.")
        else:
            print(f"ATTENTION: La feuille saisie '{nom_feuille_saisi}' n'a pas été trouvée.")

    # Nettoyage final de la feuille par défaut
    if sheet_to_delete_if_unused and sheet_to_delete_if_unused in wb_destination.sheetnames and not processed_sheet_names_in_dest:
        del wb_destination[sheet_to_delete_if_unused]
        print(f"Nettoyage final: feuille par défaut '{sheet_to_delete_if_unused}' supprimée car aucune feuille n'a été traitée.")

    if not processed_sheet_names_in_dest:
        print("Aucune feuille n'a été traitée ou copiée.")
        wb_source.close()
        wb_source_values.close()
        return None

    # Activer la première feuille traitée
    if wb_destination.sheetnames and processed_sheet_names_in_dest[0] in wb_destination.sheetnames:
        wb_destination.active = wb_destination[processed_sheet_names_in_dest[0]]

    wb_source.close() # Fermer les classeurs source
    wb_source_values.close()

    # Sauvegarder le classeur modifié en mémoire (bytes)
    try:
        # Forcer le recalcul automatique des formules à l'ouverture du fichier
        wb_destination.calculation.calcMode = 'auto'
        wb_destination.calculation.fullCalcOnLoad = True
        
        # Marquer TOUTES les feuilles comme nécessitant un recalcul complet
        for ws in wb_destination.worksheets:
            # Forcer le recalcul de toute la feuille
            ws.calcMode = 'auto'
            ws.fullCalcOnLoad = True
            # Marquer la feuille comme "dirty" pour forcer le recalcul
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':  # Si c'est une formule
                        # Forcer le recalcul en "touchant" la formule
                        formula_temp = cell.value
                        cell.value = formula_temp
        
        output_io = io.BytesIO()
        wb_destination.save(output_io)
        output_io.seek(0)
        print(f"Classeur de destination préparé en mémoire ({len(processed_sheet_names_in_dest)} feuilles) avec recalcul automatique renforcé.")
        return output_io
    except Exception as e:
        print(f"Erreur lors de la sauvegarde du fichier de destination en mémoire: {e}")
        return None

# --- Endpoints Flask ---
@app.route('/')
def health_check():
    """Health check endpoint for Railway/Render with HTML response for browsers"""
    user_agent = request.headers.get('User-Agent', '').lower()
    
    # If it's a monitoring system or API client, return JSON
    if 'go-http-client' in user_agent or request.headers.get('Accept', '').startswith('application/json'):
        return jsonify({"status": "healthy", "service": "Excel Processing API"}), 200
    
    # For browsers, return a more user-friendly HTML page
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Excel Processing API</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 0; padding: 20px; line-height: 1.6; }
            .container { max-width: 800px; margin: 0 auto; }
            h1 { color: #333; }
            .endpoint { background: #f4f4f4; padding: 10px; border-radius: 4px; margin-bottom: 10px; }
            .endpoint h3 { margin-top: 0; }
            code { background: #eee; padding: 2px 5px; border-radius: 3px; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>Excel Processing API</h1>
            <p>Status: <strong style="color: green;">Healthy</strong></p>
            
            <h2>Available Endpoints:</h2>
            
            <div class="endpoint">
                <h3>/get-sheet-names</h3>
                <p>POST request to get sheet names from an Excel file</p>
            </div>
            
            <div class="endpoint">
                <h3>/process-excel</h3>
                <p>POST request to process Excel files</p>
            </div>
            
            <div class="endpoint">
                <h3>/combine-armatures</h3>
                <p>POST request to combine armature CSV files</p>
            </div>
            
            <div class="endpoint">
                <h3>/estim-batiment</h3>
                <p>POST request to process building estimation files</p>
            </div>
            
            <p>Deployed on <a href="https://bak-5tqz.onrender.com">Render</a></p>
        </div>
    </body>
    </html>
    """
    return html, 200, {'Content-Type': 'text/html'}

@app.route('/redirect')
def redirect_to_render():
    """Redirect to the Render deployment"""
    from flask import redirect
    return redirect("https://bak-5tqz.onrender.com", code=302)

@app.route('/get-sheet-names', methods=['POST'])
def get_sheet_names_route():
    if 'excel_file' not in request.files:
        return jsonify({"error": "Aucun fichier ('excel_file') envoyé"}), 400
    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({"error": "Aucun fichier sélectionné"}), 400
    if not (file.filename.endswith('.xlsx')): # Openpyxl supporte .xlsx
        return jsonify({"error": "Type de fichier invalide. Seul .xlsx est supporté par ce backend."}), 400
    
    print(f"Fichier reçu pour extraction de noms: {file.filename}")
    file_bytes = file.read()
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        sheet_names_list = workbook.sheetnames
        workbook.close() # Important de fermer le classeur après lecture des noms
        if sheet_names_list:
            return jsonify({"sheet_names": sheet_names_list}), 200
        else:
            return jsonify({"error": "Aucun nom de feuille trouvé ou fichier vide."}), 500
    except InvalidFileException:
        return jsonify({"error": "Fichier Excel invalide ou corrompu."}), 400
    except Exception as e:
        print(f"Erreur lors de l'extraction des noms de feuilles: {e}")
        return jsonify({"error": "Erreur serveur lors de l'extraction des noms de feuilles."}), 500


@app.route('/process-excel', methods=['POST'])
def process_excel_file_route():
    if 'excel_file' not in request.files:
        return jsonify({"error": "Aucun fichier ('excel_file') envoyé"}), 400
    file = request.files['excel_file']
    sheet_names_str = request.form.get('sheet_names') # Récupérer les noms des feuilles du formulaire

    if file.filename == '': return jsonify({"error": "Aucun fichier sélectionné"}), 400
    if not sheet_names_str: return jsonify({"error": "Noms de feuilles à traiter non fournis ('sheet_names')"}), 400
    if not (file.filename.endswith('.xlsx')):
        return jsonify({"error": "Type de fichier invalide. Seul .xlsx est supporté."}), 400

    print(f"Fichier reçu pour traitement: {file.filename}, Feuilles: {sheet_names_str}")
    file_bytes = file.read()
    
    processed_file_io = traiter_fichier_excel_core(file_bytes, sheet_names_str) 

    if processed_file_io:
        print(f"Envoi du fichier traité '{file.filename}'")
        return send_file(
            processed_file_io,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=False,  # Changé à False pour ouvrir directement
        download_name=f"traite_{file.filename}"
        )
    else:
        return jsonify({"error": "Erreur serveur lors du traitement du fichier Excel."}), 500

@app.route('/combine-armatures', methods=['POST'])
def combine_armatures_route():
    if not request.files:
        print("Aucun fichier n'a été envoyé pour la combinaison d'armatures.")
        return jsonify({"error": "Aucun fichier envoyé."}), 400

    files_data = []
    # Utiliser getlist pour récupérer TOUS les fichiers avec la même clé
    uploaded_files = request.files.getlist('csv_files')
    print(f"Nombre de fichiers reçus via getlist('csv_files'): {len(uploaded_files)}")
    
    for file in uploaded_files:
        if file and file.filename:
            print(f"Fichier reçu pour combinaison: {file.filename}")
            files_data.append({
                'name': file.filename,
                'bytes': file.read()
            })
    
    if not files_data:
        print("La liste des fichiers pour combinaison est vide après traitement initial.")
        return jsonify({"error": "Aucuns fichiers valides trouvés dans la requête."}), 400

    print(f"{len(files_data)} fichier(s) prêt(s) pour la fonction process_armature_csvs.")
    
    try:
        output_excel_io, output_filename = process_armature_csvs(files_data)
        
        if output_excel_io and output_filename:
            print(f"Envoi du fichier combiné d'armatures: {output_filename}")
            return send_file(
                output_excel_io,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True, # True pour forcer le téléchargement
                download_name=output_filename
            )
        else:
            print("process_armature_csvs n'a pas retourné de fichier valide.")
            # output_filename pourrait contenir un message d'erreur ou un nom de fichier d'erreur
            error_message = f"Erreur lors de la combinaison des armatures. Détail: {output_filename if output_filename else 'Inconnu'}"
            return jsonify({"error": error_message}), 500
    except Exception as e:
        print(f"Exception lors de l'appel à process_armature_csvs ou de l'envoi du fichier: {e}")
        return jsonify({"error": f"Erreur serveur critique lors de la combinaison des armatures: {str(e)}"}), 500

@app.route('/estim-batiment', methods=['POST'])
def estim_batiment_route():
    """
    Endpoint pour traiter les fichiers Excel d'estimation de bâtiment.
    Accepte un fichier Excel avec les feuilles requises et génère un devis détaillé.
    """
    if not request.files:
        print("Aucun fichier n'a été envoyé pour l'estimation bâtiment.")
        return jsonify({"error": "Aucun fichier envoyé."}), 400

    if 'excel_file' not in request.files:
        print("Clé 'excel_file' manquante dans les fichiers envoyés.")
        return jsonify({"error": "Fichier Excel requis avec la clé 'excel_file'."}), 400

    uploaded_file = request.files['excel_file']
    
    if not uploaded_file or not uploaded_file.filename:
        print("Fichier Excel vide ou sans nom.")
        return jsonify({"error": "Fichier Excel valide requis."}), 400

    print(f"Fichier reçu pour estimation bâtiment: {uploaded_file.filename}")

    try:
        # Lire le contenu du fichier
        file_bytes = uploaded_file.read()
        
        if not file_bytes:
            print("Le fichier reçu est vide.")
            return jsonify({"error": "Le fichier envoyé est vide."}), 400

        print(f"Taille du fichier reçu: {len(file_bytes)} bytes")
        
        # Traiter le fichier avec la fonction EstimBatiment
        output_excel_io, output_filename = process_estim_batiment(file_bytes)
        
        if output_excel_io and output_filename:
            print(f"Envoi du fichier d'estimation: {output_filename}")
            return send_file(
                output_excel_io,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=output_filename
            )
        else:
            # output_filename contient le message d'erreur
            error_message = output_filename if output_filename else "Erreur inconnue lors du traitement"
            print(f"Erreur lors du traitement EstimBatiment: {error_message}")
            return jsonify({"error": error_message}), 500
            
    except Exception as e:
        print(f"Exception lors du traitement EstimBatiment: {e}")
        return jsonify({"error": f"Erreur serveur critique lors du traitement EstimBatiment: {str(e)}"}), 500

# --- Fonction de traitement EstimBatiment ---
def process_estim_batiment(excel_file_bytes):
    """
    Traite un fichier Excel d'estimation et génère un devis détaillé.
    Basé sur la logique de main.py du module EstimBatiment.
    
    Args:
        excel_file_bytes: Bytes du fichier Excel d'entrée
        
    Returns:
        tuple: (output_excel_io, output_filename) ou (None, error_message)
    """
    try:
        print("Traitement EstimBatiment - Chargement du classeur...")
        
        # Charge le classeur une première fois pour accéder aux formules (data_only=False)
        input_wb_formulas = openpyxl.load_workbook(io.BytesIO(excel_file_bytes), data_only=False)
        
        # Charge le classeur une deuxième fois pour obtenir les valeurs calculées (data_only=True)
        input_wb_values = openpyxl.load_workbook(io.BytesIO(excel_file_bytes), data_only=True)

    except Exception as e:
        print(f"Erreur lors de l'ouverture du fichier d'estimation: {e}")
        return None, f"Erreur lors de l'ouverture du fichier: {str(e)}"

    # Vérifie la présence des feuilles nécessaires
    required_formula_sheets = ["qt", "calcul", "Peinture", "Revetement", "Toiture"]
    required_value_sheets = ["open", "Electricite", "Plomberie"]

    sheets_formulas = {}
    sheets_values = {}
    
    for sheet_name in required_formula_sheets:
        if sheet_name not in input_wb_formulas.sheetnames:
            print(f"La feuille '{sheet_name}' est manquante dans le fichier d'entrée (pour les formules).")
            sheets_formulas[sheet_name] = None
        else:
            sheets_formulas[sheet_name] = input_wb_formulas[sheet_name]

    for sheet_name in required_value_sheets:
        if sheet_name not in input_wb_values.sheetnames:
            print(f"La feuille '{sheet_name}' est manquante dans le fichier d'entrée (pour les valeurs).")
            sheets_values[sheet_name] = None
        else:
            sheets_values[sheet_name] = input_wb_values[sheet_name]

    # Assign sheets to variables for clarity
    qt_sheet = sheets_formulas["qt"]
    calcul_sheet = sheets_formulas["calcul"]
    open_sheet = sheets_values["open"]
    electricite_sheet = sheets_values["Electricite"]
    plomberie_sheet = sheets_values["Plomberie"]
    peinture_sheet = sheets_formulas["Peinture"]
    revetement_sheet = sheets_formulas["Revetement"]
    toiture_sheet = sheets_formulas["Toiture"]

    # Vérification des feuilles critiques
    if qt_sheet is None:
        return None, "La feuille 'qt' est obligatoire et manquante dans le fichier."
    if calcul_sheet is None:
        return None, "La feuille 'calcul' est obligatoire et manquante dans le fichier."

    # --- Lecture des données ---
    print("Lecture des données de la feuille 'qt'...")
    qt_data_dict = get_qt_data(qt_sheet)
    if not qt_data_dict:
        print("AVERTISSEMENT: Aucune donnée lue depuis la feuille 'qt'.")
    
    open_data_list = []
    if open_sheet:
        print("Lecture des données de la feuille 'open'...")
        open_data_list = get_open_data(open_sheet)

    electricite_data_list = []
    if electricite_sheet:
        print("Lecture des données de la feuille 'Electricite'...")
        electricite_data_list = get_simple_block_data(electricite_sheet)

    plomberie_data_list = []
    if plomberie_sheet:
        print("Lecture des données de la feuille 'Plomberie'...")
        plomberie_data_list = get_simple_block_data(plomberie_sheet)

    peinture_data_list = []
    if peinture_sheet:
        print("Lecture des données de la feuille 'Peinture'...")
        peinture_data_list = get_formula_block_data(peinture_sheet)

    revetement_data_list = []
    if revetement_sheet:
        print("Lecture des données de la feuille 'Revetement'...")
        revetement_data_list = get_formula_block_data(revetement_sheet)

    toiture_data_list = []
    if toiture_sheet:
        print("Lecture des données de la feuille 'Toiture'...")
        toiture_data_list = get_formula_block_data(toiture_sheet)

    # --- Configuration du classeur de sortie ---
    output_wb = openpyxl.Workbook()
    if "Sheet" in output_wb.sheetnames:
        main_output_sheet = output_wb["Sheet"]
        main_output_sheet.title = "Estimation Globale"
    else:
        main_output_sheet = output_wb.create_sheet("Estimation Globale", 0) 

    # --- Liste pour le récapitulatif ---
    recap_entries = []

    # --- Traitement et écriture des blocs ---
    current_excel_row = 1 

    print("Analyse de la feuille 'calcul' et génération des tableaux...")
    current_excel_row = parse_calcul_sheet_and_process_blocks(calcul_sheet, qt_data_dict, main_output_sheet, recap_entries)

    # Bloc IV: Menuiserie
    if open_data_list:
        print("Traitement du bloc IV: Menuiserie...")
        current_excel_row = process_menuiserie_block(open_data_list, main_output_sheet, current_excel_row, recap_entries)

    # Bloc V: Electricité
    if electricite_data_list:
        print("Traitement du bloc V: Electricité...")
        current_excel_row = process_simple_block(electricite_data_list, main_output_sheet, current_excel_row, "V", "ELECTRICITE", 1, recap_entries)

    # Bloc VI: Plomberie
    if plomberie_data_list:
        print("Traitement du bloc VI: Plomberie...")
        current_excel_row = process_simple_block(plomberie_data_list, main_output_sheet, current_excel_row, "VI", "PLOMBERIE SANITAIRE", 1, recap_entries)

    # Bloc VII: Revetement
    if revetement_data_list:
        print("Traitement du bloc VII: Revetement...")
        current_excel_row = process_formula_block(revetement_data_list, qt_data_dict, main_output_sheet, current_excel_row, "VII", "REVETEMENT", 1, recap_entries)

    # Bloc VIII: Peinture
    if peinture_data_list:
        print("Traitement du bloc VIII: Peinture...")
        current_excel_row = process_formula_block(peinture_data_list, qt_data_dict, main_output_sheet, current_excel_row, "VIII", "PEINTURE", 1, recap_entries)

    # Bloc IX: Toiture
    if toiture_data_list:
        print("Traitement du bloc IX: Toiture...")
        current_excel_row = process_formula_block(toiture_data_list, qt_data_dict, main_output_sheet, current_excel_row, "IX", "TOITURE", 1, recap_entries)

    # --- Ajout du récapitulatif ---
    if recap_entries:
        print("Génération du bloc RÉCAPITULATIF...")
        current_excel_row = write_recap_block(main_output_sheet, current_excel_row, recap_entries)

    # --- Vérification finale et sauvegarde ---
    if main_output_sheet.max_row <= 1: 
        return None, "Aucun bloc n'a été traité ou aucune donnée valide trouvée."

    # Génération du nom de fichier de sortie
    output_filename = "Estimation_Batiment_Calculee.xlsx"

    # Sauvegarde en mémoire
    try:
        output_io = io.BytesIO()
        output_wb.save(output_io)
        output_io.seek(0)
        print("Fichier d'estimation généré avec succès.")
        return output_io, output_filename
    except Exception as e:
        print(f"Erreur lors de la sauvegarde du fichier d'estimation: {e}")
        return None, f"Erreur lors de la sauvegarde: {str(e)}"

# --- Lancement de l'application ---
if __name__ == '__main__':
    print("Démarrage du serveur Flask pour traitement Excel...")
    # Configuration pour déploiement (Heroku, etc.)
    import os
    port = int(os.environ.get('PORT', 5000))
    host = '0.0.0.0'
    debug = os.environ.get('FLASK_ENV') != 'production'
    
    app.run(debug=debug, host=host, port=port)