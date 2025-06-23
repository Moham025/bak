# main.py
import tkinter as tk
from tkinter import filedialog
import os
import openpyxl

# Import des fonctions des autres modules
from data_reader import get_qt_data, get_open_data, get_simple_block_data, get_formula_block_data
from calculation_engine import parse_calcul_sheet_and_process_blocks, process_menuiserie_block, process_simple_block, process_formula_block, write_recap_block

def main():
    """
    Fonction principale de l'application.
    Gère le flux de travail : sélection de fichiers, lecture, traitement et sauvegarde.
    """
    root = tk.Tk()
    root.withdraw() # Cache la fenêtre principale de Tkinter

    # Demande à l'utilisateur de sélectionner le fichier Excel d'entrée
    input_filepath = filedialog.askopenfilename(
        title="Sélectionnez le fichier Excel d'estimation (avec toutes les feuilles)",
        filetypes=(("Fichiers Excel", "*.xlsx *.xls"), ("Tous les fichiers", "*.*"))
    )
    if not input_filepath:
        print("Aucun fichier sélectionné. Annulation.")
        return

    try:
        print(f"Chargement du classeur d'entrée pour formules (calcul, qt, Peinture, Revetement, Toiture): {input_filepath}")
        # Charge le classeur une première fois pour accéder aux formules (data_only=False)
        input_wb_formulas = openpyxl.load_workbook(input_filepath, data_only=False)
        
        print(f"Chargement du classeur d'entrée pour valeurs (open, Electricite, Plomberie): {input_filepath}")
        # Charge le classeur une deuxième fois pour obtenir les valeurs calculées (data_only=True)
        input_wb_values = openpyxl.load_workbook(input_filepath, data_only=True)

    except Exception as e:
        print(f"Erreur lors de l'ouverture du fichier d'entrée: {e}")
        return

    # Vérifie la présence des feuilles nécessaires dans les classeurs appropriés
    required_formula_sheets = ["qt", "calcul", "Peinture", "Revetement", "Toiture"]
    required_value_sheets = ["open", "Electricite", "Plomberie"]

    sheets_formulas = {}
    sheets_values = {}
    
    for sheet_name in required_formula_sheets:
        if sheet_name not in input_wb_formulas.sheetnames:
            print(f"La feuille '{sheet_name}' est manquante dans le fichier d'entrée (pour les formules).")
            sheets_formulas[sheet_name] = None
        else:
            sheets_formulas[sheet_name] = input_wb_formulas[sheet_name]

    for sheet_name in required_value_sheets:
        if sheet_name not in input_wb_values.sheetnames:
            print(f"La feuille '{sheet_name}' est manquante dans le fichier d'entrée (pour les valeurs).")
            sheets_values[sheet_name] = None
        else:
            sheets_values[sheet_name] = input_wb_values[sheet_name]

    # Assign sheets to variables for clarity
    qt_sheet = sheets_formulas["qt"]
    calcul_sheet = sheets_formulas["calcul"]
    open_sheet = sheets_values["open"]
    electricite_sheet = sheets_values["Electricite"]
    plomberie_sheet = sheets_values["Plomberie"]
    peinture_sheet = sheets_formulas["Peinture"]
    revetement_sheet = sheets_formulas["Revetement"]
    toiture_sheet = sheets_formulas["Toiture"]

    # --- Lecture des données ---
    print("Lecture des données de la feuille 'qt'...")
    qt_data_dict = get_qt_data(qt_sheet)
    if not qt_data_dict:
        print("AVERTISSEMENT: Aucune donnée lue depuis la feuille 'qt'. Les calculs de quantité échoueront probablement.")
    
    open_data_list = []
    if open_sheet:
        print("Lecture des données de la feuille 'open'...")
        open_data_list = get_open_data(open_sheet)
        if not open_data_list:
            print("AVERTISSEMENT: Aucune donnée valide lue depuis la feuille 'open'.")

    electricite_data_list = []
    if electricite_sheet:
        print("Lecture des données de la feuille 'Electricite'...")
        electricite_data_list = get_simple_block_data(electricite_sheet)
        if not electricite_data_list:
            print("AVERTISSEMENT: Aucune donnée valide lue depuis la feuille 'Electricite'.")

    plomberie_data_list = []
    if plomberie_sheet:
        print("Lecture des données de la feuille 'Plomberie'...")
        plomberie_data_list = get_simple_block_data(plomberie_sheet)
        if not plomberie_data_list:
            print("AVERTISSEMENT: Aucune donnée valide lue depuis la feuille 'Plomberie'.")

    peinture_data_list = []
    if peinture_sheet:
        print("Lecture des données de la feuille 'Peinture'...")
        peinture_data_list = get_formula_block_data(peinture_sheet)
        if not peinture_data_list:
            print("AVERTISSEMENT: Aucune donnée valide lue depuis la feuille 'Peinture'.")

    revetement_data_list = []
    if revetement_sheet:
        print("Lecture des données de la feuille 'Revetement'...")
        revetement_data_list = get_formula_block_data(revetement_sheet)
        if not revetement_data_list:
            print("AVERTISSEMENT: Aucune donnée valide lue depuis la feuille 'Revetement'.")

    toiture_data_list = []
    if toiture_sheet:
        print("Lecture des données de la feuille 'Toiture'...")
        toiture_data_list = get_formula_block_data(toiture_sheet)
        if not toiture_data_list:
            print("AVERTISSEMENT: Aucune donnée valide lue depuis la feuille 'Toiture'.")


    # --- Configuration du classeur de sortie ---
    output_wb = openpyxl.Workbook()
    if "Sheet" in output_wb.sheetnames:
        main_output_sheet = output_wb["Sheet"]
        main_output_sheet.title = "Estimation Globale"
    else:
        main_output_sheet = output_wb.create_sheet("Estimation Globale", 0) 

    # --- Liste pour le récapitulatif ---
    recap_entries = [] # Chaque entrée: {'roman', 'title', 'total_cell_ref', 'numeric_total'}

    # --- Traitement et écriture des blocs ---
    current_excel_row = 1 

    print("\nAnalyse de la feuille 'calcul' (Blocs I, II, III) et génération des tableaux...")
    current_excel_row = parse_calcul_sheet_and_process_blocks(calcul_sheet, qt_data_dict, main_output_sheet, recap_entries)

    # Bloc IV: Menuiserie
    if open_data_list:
        print("\nTraitement du bloc IV: Menuiserie...")
        current_excel_row = process_menuiserie_block(open_data_list, main_output_sheet, current_excel_row, recap_entries)
    else:
        print("\nAucune donnée pour le bloc Menuiserie trouvée ou feuille 'open' manquante, le bloc ne sera pas ajouté.")

    # Bloc V: Electricité
    if electricite_data_list:
        print("\nTraitement du bloc V: Electricité...")
        # L'item_start_num est 1 car c'est V.1, V.2 etc.
        current_excel_row = process_simple_block(electricite_data_list, main_output_sheet, current_excel_row, "V", "ELECTRICITE", 1, recap_entries)
    else:
        print("\nAucune donnée pour le bloc Electricité trouvée ou feuille 'Electricite' manquante, le bloc ne sera pas ajouté.")

    # Bloc VI: Plomberie
    if plomberie_data_list:
        print("\nTraitement du bloc VI: Plomberie...")
        current_excel_row = process_simple_block(plomberie_data_list, main_output_sheet, current_excel_row, "VI", "PLOMBERIE SANITAIRE", 1, recap_entries)
    else:
        print("\nAucune donnée pour le bloc Plomberie trouvée ou feuille 'Plomberie' manquante, le bloc ne sera pas ajouté.")

    # Bloc VII: Revetement
    if revetement_data_list:
        print("\nTraitement du bloc VII: Revetement...")
        current_excel_row = process_formula_block(revetement_data_list, qt_data_dict, main_output_sheet, current_excel_row, "VII", "REVETEMENT", 1, recap_entries)
    else:
        print("\nAucune donnée pour le bloc Revetement trouvée ou feuille 'Revetement' manquante, le bloc ne sera pas ajouté.")

    # Bloc VIII: Peinture
    if peinture_data_list:
        print("\nTraitement du bloc VIII: Peinture...")
        current_excel_row = process_formula_block(peinture_data_list, qt_data_dict, main_output_sheet, current_excel_row, "VIII", "PEINTURE", 1, recap_entries)
    else:
        print("\nAucune donnée pour le bloc Peinture trouvée ou feuille 'Peinture' manquante, le bloc ne sera pas ajouté.")

    # Bloc IX: Toiture
    if toiture_data_list:
        print("\nTraitement du bloc IX: Toiture...")
        current_excel_row = process_formula_block(toiture_data_list, qt_data_dict, main_output_sheet, current_excel_row, "IX", "TOITURE", 1, recap_entries)
    else:
        print("\nAucune donnée pour le bloc Toiture trouvée ou feuille 'Toiture' manquante, le bloc ne sera pas ajouté.")


    # --- Ajout du récapitulatif ---
    if recap_entries: # Si des blocs ont été traités et ajoutés au récapitulatif
        print("\nGénération du bloc RÉCAPITULATIF...")
        current_excel_row = write_recap_block(main_output_sheet, current_excel_row, recap_entries)
    else:
        print("\nAucune donnée de récapitulatif à générer.")


    # --- Vérification finale et sauvegarde ---
    if main_output_sheet.max_row <= 1: 
        print("\nAucun bloc n'a été traité ou aucune donnée valide trouvée. Aucun fichier de sortie généré.")
        return

    base, ext = os.path.splitext(os.path.basename(input_filepath))
    output_filename = f"{base}_estimation_globale_calculee.xlsx"
    
    output_filepath = filedialog.asksaveasfilename(
        title="Enregistrer le fichier d'estimation calculée sous...",
        defaultextension=".xlsx",
        initialfile=output_filename,
        filetypes=(("Fichiers Excel", "*.xlsx"), ("Tous les fichiers", "*.*"))
    )

    if output_filepath:
        try:
            output_wb.save(output_filepath)
            print(f"\nClasseur d'estimation calculée sauvegardé avec succès sous: {output_filepath}")
        except PermissionError:
            print(f"Erreur de permission : Impossible de sauvegarder '{output_filepath}'. Vérifiez qu'il n'est pas ouvert.")
        except Exception as e:
            print(f"Une erreur est survenue lors de la sauvegarde : {e}")
    else:
        print("\nSauvegarde annulée par l'utilisateur.")

if __name__ == "__main__":
    main()
