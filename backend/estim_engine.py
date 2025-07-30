# estim_engine.py
import openpyxl
import io

# Import des modules de traitement depuis le sous-dossier "EstimBatiment"
from EstimBatiment.data_reader import get_qt_data, get_open_data, get_simple_block_data, get_formula_block_data
from EstimBatiment.calculation_engine import parse_calcul_sheet_and_process_blocks, process_menuiserie_block, process_simple_block, process_formula_block, write_recap_block

def process_estim_batiment(excel_file_bytes):
    """
    Traite un fichier Excel d'estimation et génère un devis détaillé.
    C'est la logique métier principale, sans code web.
    
    Args:
        excel_file_bytes: Bytes du fichier Excel d'entrée
        
    Returns:
        tuple: (output_excel_io, output_filename) ou (None, error_message)
    """
    try:
        print("Traitement EstimBatiment - Chargement du classeur...")
        input_wb_formulas = openpyxl.load_workbook(io.BytesIO(excel_file_bytes), data_only=False)
        input_wb_values = openpyxl.load_workbook(io.BytesIO(excel_file_bytes), data_only=True)
    except Exception as e:
        print(f"Erreur lors de l'ouverture du fichier d'estimation: {e}")
        return None, f"Erreur lors de l'ouverture du fichier: {str(e)}"

    # La logique de lecture des feuilles est maintenant correcte et centralisée ici
    required_formula_sheets = ["calcul", "Peinture", "Revetement", "Toiture"]
    required_value_sheets = ["qt", "open", "Electricite", "Plomberie"]

    sheets_formulas = {name: input_wb_formulas.get(name) for name in required_formula_sheets}
    sheets_values = {name: input_wb_values.get(name) for name in required_value_sheets}

    qt_sheet = sheets_values.get("qt")
    calcul_sheet = sheets_formulas.get("calcul")
    
    if qt_sheet is None:
        return None, "La feuille 'qt' est obligatoire et manquante dans le fichier."
    if calcul_sheet is None:
        return None, "La feuille 'calcul' est obligatoire et manquante dans le fichier."

    # --- Lecture des données ---
    print("Lecture des données...")
    qt_data_dict = get_qt_data(qt_sheet)
    open_data_list = get_open_data(sheets_values["open"]) if sheets_values.get("open") else []
    electricite_data_list = get_simple_block_data(sheets_values["Electricite"]) if sheets_values.get("Electricite") else []
    plomberie_data_list = get_simple_block_data(sheets_values["Plomberie"]) if sheets_values.get("Plomberie") else []
    peinture_data_list = get_formula_block_data(sheets_formulas["Peinture"]) if sheets_formulas.get("Peinture") else []
    revetement_data_list = get_formula_block_data(sheets_formulas["Revetement"]) if sheets_formulas.get("Revetement") else []
    toiture_data_list = get_formula_block_data(sheets_formulas["Toiture"]) if sheets_formulas.get("Toiture") else []

    # --- Configuration et traitement du classeur de sortie ---
    output_wb = openpyxl.Workbook()
    main_output_sheet = output_wb.active
    main_output_sheet.title = "Estimation Globale"
    
    recap_entries = []
    current_excel_row = 1 

    print("Analyse de la feuille 'calcul' et génération des tableaux...")
    current_excel_row = parse_calcul_sheet_and_process_blocks(calcul_sheet, qt_data_dict, main_output_sheet, recap_entries)

    if open_data_list:
        current_excel_row = process_menuiserie_block(open_data_list, main_output_sheet, current_excel_row, recap_entries)
    if electricite_data_list:
        current_excel_row = process_simple_block(electricite_data_list, main_output_sheet, current_excel_row, "V", "ELECTRICITE", 1, recap_entries)
    if plomberie_data_list:
        current_excel_row = process_simple_block(plomberie_data_list, main_output_sheet, current_excel_row, "VI", "PLOMBERIE SANITAIRE", 1, recap_entries)
    if revetement_data_list:
        current_excel_row = process_formula_block(revetement_data_list, qt_data_dict, main_output_sheet, current_excel_row, "VII", "REVETEMENT", 1, recap_entries)
    if peinture_data_list:
        current_excel_row = process_formula_block(peinture_data_list, qt_data_dict, main_output_sheet, current_excel_row, "VIII", "PEINTURE", 1, recap_entries)
    if toiture_data_list:
        current_excel_row = process_formula_block(toiture_data_list, qt_data_dict, main_output_sheet, current_excel_row, "IX", "TOITURE", 1, recap_entries)

    if recap_entries:
        write_recap_block(main_output_sheet, current_excel_row, recap_entries)

    if main_output_sheet.max_row <= 1: 
        return None, "Aucun bloc n'a été traité ou aucune donnée valide trouvée."

    # --- Sauvegarde en mémoire ---
    try:
        output_filename = "Estimation_Batiment_Calculee.xlsx"
        output_io = io.BytesIO()
        output_wb.save(output_io)
        output_io.seek(0)
        print("Fichier d'estimation généré avec succès.")
        return output_io, output_filename
    except Exception as e:
        print(f"Erreur lors de la sauvegarde du fichier d'estimation: {e}")
        return None, f"Erreur lors de la sauvegarde: {str(e)}"
