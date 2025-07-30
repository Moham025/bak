# utility_routes.py
from flask import Blueprint, request, jsonify, send_file
import openpyxl
import io
import re
from copy import copy
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils.cell import get_column_letter

# Import des fonctions de conversion de nombre en lettre
from covnumletter import conv_number_letter as cl_conv_number_letter

# --- Blueprint Setup ---
utility_bp = Blueprint('utility', __name__)

# --- Helper Functions (from bon_a_envoye.py) ---

def trouver_nom_feuille_original(nom_saisi, noms_feuilles_sources_dict):
    nom_normalise_saisi = nom_saisi.strip().lower()
    return noms_feuilles_sources_dict.get(nom_normalise_saisi, "")

def est_une_feuille_recap(nom_feuille_original):
    nom_lower = nom_feuille_original.strip().lower()
    recap_keywords = ["recap", "récap", "summary", "synthese", "synthèse"]
    specific_recap_names = ["recapitulatif", "récapitulatif"]
    return any(keyword in nom_lower for keyword in recap_keywords) or nom_lower in specific_recap_names

def add_quotes_if_necessary(sheet_name):
    if re.search(r"[\s!@#$%^&*()+={}\[\]:;\"'<>,.?/\\|-]", sheet_name):
        return f"'{sheet_name.replace('\'', '\'\'')}'"
    return sheet_name

def copier_feuille_manuellement(ws_source, wb_destination, nouveau_nom_feuille):
    if nouveau_nom_feuille in wb_destination.sheetnames:
        del wb_destination[nouveau_nom_feuille]
    ws_destination = wb_destination.create_sheet(title=nouveau_nom_feuille)
    for row in ws_source.iter_rows():
        for cell in row:
            new_cell = ws_destination.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font, new_cell.border, new_cell.fill = copy(cell.font), copy(cell.border), copy(cell.fill)
                new_cell.number_format, new_cell.protection, new_cell.alignment = cell.number_format, copy(cell.protection), copy(cell.alignment)
    for col_letter, dim in ws_source.column_dimensions.items(): ws_destination.column_dimensions[col_letter].width = dim.width
    for row_num, dim in ws_source.row_dimensions.items(): ws_destination.row_dimensions[row_num].height = dim.height
    for merged_range_str in ws_source.merged_cells.ranges: ws_destination.merge_cells(str(merged_range_str))
    return ws_destination

def nettoyer_total_en_lettres(ws_copie, ws_original_source_values):
    prefixe = "Arrêter le présent devis estimatif à la somme de :"
    ligne_total_general = None
    for row in range(1, ws_copie.max_row + 1):
        cell_a = ws_copie[f"A{row}"]
        if cell_a.value and isinstance(cell_a.value, str) and "TOTAL GENERAL" in cell_a.value.upper():
            ligne_total_general = row
            break
    if ligne_total_general:
        cell_f_total_coord = f"F{ligne_total_general}"
        montant_total = ws_original_source_values[cell_f_total_coord].value
        if isinstance(montant_total, (int, float)) and montant_total > 0:
            texte_total_lettres = cl_conv_number_letter(montant_total, devise=1, langue=0)
            ws_copie[f"A{ligne_total_general + 1}"].value = f"{prefixe} {texte_total_lettres}"

def modifier_liens_externes_feuille_recap(ws_recap, wb_cible):
    for col_num in [6, ws_recap.max_column] if ws_recap.max_column != 6 else [6]:
        for r in range(1, ws_recap.max_row + 1):
            cell = ws_recap.cell(row=r, column=col_num)
            if cell.data_type == 'f':
                formula = str(cell.value)
                match = re.match(r"=.*?['\[]([^\]']+)['\]]([^!]+)!(.+)", formula, re.IGNORECASE) or \
                        re.match(r"=([^!\[]+)!(.+)", formula, re.IGNORECASE)
                if match:
                    groups = match.groups()
                    sheet_name_raw = groups[-2].strip("'")
                    cell_ref = groups[-1]
                    target_sheet = f"{sheet_name_raw}_copie"
                    if target_sheet in wb_cible.sheetnames:
                        cell.value = f"={add_quotes_if_necessary(target_sheet)}!{cell_ref}"

def traiter_fichier_excel_core(bytes_fichier_source, noms_feuilles_a_traiter_str):
    try:
        wb_source = openpyxl.load_workbook(io.BytesIO(bytes_fichier_source), data_only=False)
        wb_source_values = openpyxl.load_workbook(io.BytesIO(bytes_fichier_source), data_only=True)
    except Exception as e:
        return None, f"Impossible de charger le fichier source: {e}"

    noms_feuilles_sources_dict = {name.strip().lower(): name.strip() for name in wb_source.sheetnames}
    l_array_str = [s.strip() for s in noms_feuilles_a_traiter_str.split(',') if s.strip()]
    
    wb_destination = openpyxl.Workbook()
    if "Sheet" in wb_destination.sheetnames:
        wb_destination.remove(wb_destination["Sheet"])

    for nom_feuille_saisi in l_array_str:
        nom_feuille_source_original = trouver_nom_feuille_original(nom_feuille_saisi, noms_feuilles_sources_dict)
        if nom_feuille_source_original:
            ws_original_source = wb_source[nom_feuille_source_original]
            ws_original_source_values = wb_source_values[nom_feuille_source_original]
            nom_feuille_copie_dest = f"{nom_feuille_source_original.strip()}_copie"
            
            ws_copie_dest = copier_feuille_manuellement(ws_original_source, wb_destination, nom_feuille_copie_dest)
            
            for col in ["D", "E"]:
                for row in range(1, ws_copie_dest.max_row + 1):
                    if ws_copie_dest[f"{col}{row}"].data_type == 'f':
                        ws_copie_dest[f"{col}{row}"].value = ws_original_source_values[f"{col}{row}"].value
            
            nettoyer_total_en_lettres(ws_copie_dest, ws_original_source_values)
            ws_copie_dest.delete_cols(7, 4)
            
            if est_une_feuille_recap(nom_feuille_source_original):
                modifier_liens_externes_feuille_recap(ws_copie_dest, wb_destination)

    if not wb_destination.sheetnames:
        return None, "Aucune feuille valide n'a été traitée."

    output_io = io.BytesIO()
    wb_destination.save(output_io)
    output_io.seek(0)
    return output_io, "fichier_traite.xlsx"

# --- Routes ---

@utility_bp.route('/get-sheet-names', methods=['POST'])
def get_sheet_names_route():
    if 'excel_file' not in request.files:
        return jsonify({"error": "Aucun fichier ('excel_file') envoyé"}), 400
    file = request.files['excel_file']
    if not file.filename:
        return jsonify({"error": "Aucun fichier sélectionné"}), 400
    
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file.read()), read_only=True)
        return jsonify({"sheet_names": workbook.sheetnames}), 200
    except InvalidFileException:
        return jsonify({"error": "Fichier Excel invalide ou corrompu."}), 400
    except Exception as e:
        return jsonify({"error": f"Erreur serveur: {str(e)}"}), 500

@utility_bp.route('/process-excel', methods=['POST'])
def process_excel_file_route():
    if 'excel_file' not in request.files:
        return jsonify({"error": "Aucun fichier ('excel_file') envoyé"}), 400
    file = request.files['excel_file']
    sheet_names_str = request.form.get('sheet_names')

    if not sheet_names_str:
        return jsonify({"error": "Noms de feuilles à traiter non fournis ('sheet_names')"}), 400

    processed_file_io, output_filename_or_error = traiter_fichier_excel_core(file.read(), sheet_names_str) 

    if processed_file_io:
        return send_file(
            processed_file_io,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=output_filename_or_error
        )
    else:
        return jsonify({"error": output_filename_or_error or "Erreur serveur lors du traitement."}), 500
