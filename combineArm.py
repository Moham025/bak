import pandas as pd
import io
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv

# --- Constantes (inspirées du VBA) ---
PRIX_TONNE = 525000
# Nbr barre/tonne pour HA6, HA8, HA10, HA12, HA14
NBR_BARRE_TONNE_MAP = {
    6: 375,
    8: 210,
    10: 135,
    12: 93,
    14: 68
}
HA_TYPES = [6, 8, 10, 12, 14] # Types d'armatures à traiter

def clean_value_for_numeric_conversion(value):
    """Nettoie les préfixes et convertit la virgule décimale en point.
       Prépare la valeur pour une conversion en type numérique.
    """
    if pd.isna(value):
        return None
    s_value = str(value).strip()
    # Supprimer les préfixes (amélioration : plus de préfixes)
    s_value = s_value.replace("A = ", "").replace("B = ", "").replace("C = ", "")
    s_value = s_value.replace("A=", "").replace("B=", "").replace("C=", "")
    s_value = s_value.replace("A ", "").replace("B ", "").replace("C ", "")
    # Remplacer la virgule décimale par un point
    s_value = s_value.replace(",", ".")
    # Si après nettoyage on a une chaîne vide, retourner None
    if not s_value.strip():
        return None
    return s_value

def process_armature_csvs(csv_file_contents_list):
    """
    Fonction principale pour traiter une liste de contenus de fichiers CSV.
    Chaque item de csv_file_contents_list est un dictionnaire 
    comme {'name': 'filename.csv', 'bytes': b'contenu_csv'}
    """
    print(f"=== DÉBUT TRAITEMENT CSV ===")
    print(f"Nombre de fichiers reçus: {len(csv_file_contents_list)}")
    for i, file_info in enumerate(csv_file_contents_list):
        file_name = file_info['name']
        file_size = len(file_info['bytes'])
        print(f"  Fichier {i+1}: {file_name} - Taille: {file_size} bytes")
    print(f"=== DÉBUT LECTURE ET COMBINAISON ===")
    
    all_data_frames = []
    column_names_from_first_csv = None

    # 1. Lire et Combiner tous les fichiers CSV
    for i, file_info in enumerate(csv_file_contents_list):
        file_bytes = file_info['bytes']
        file_name = file_info['name']
        print(f"Traitement du fichier: {file_name}...")
        content_str = None
        encodings_to_try = ['utf-8-sig', 'latin1', 'cp1252'] # Ordre des essais
        
        for encoding in encodings_to_try:
            try:
                content_str = file_bytes.decode(encoding)
                print(f"  Fichier {file_name} décodé avec succès en utilisant {encoding}.")
                break # Sortir de la boucle si le décodage réussit
            except UnicodeDecodeError:
                print(f"  Échec du décodage de {file_name} avec {encoding}.")
                continue # Essayer l\'encodage suivant
        
        if content_str is None:
            print(f"Erreur: Impossible de décoder le fichier {file_name} avec les encodages testés.")
            # Selon la robustesse désirée, on pourrait sauter le fichier ou lever une erreur.
            continue

        try:
            # Afficher les premières lignes brutes pour diagnostic
            first_lines = content_str.split('\n')[:5]  # Prendre plus de lignes pour diagnostic
            print(f"  >>> Premières lignes brutes de {file_name}:")
            for j, line in enumerate(first_lines):
                print(f"    Ligne {j+1}: {repr(line[:100])}")  # Limiter à 100 chars pour lisibilité
            
            # Nettoyer le contenu en supprimant les lignes vides au début
            lines = content_str.split('\n')
            non_empty_lines = []
            for line in lines:
                if line.strip() and not all(c in ';\t, ' for c in line):  # Ignorer les lignes vides ou avec que des séparateurs
                    non_empty_lines.append(line)
            
            if not non_empty_lines:
                print(f"  >>> Fichier {file_name} ne contient aucune ligne de données valide.")
                continue
                
            cleaned_content = '\n'.join(non_empty_lines)
            print(f"  >>> Fichier {file_name}: {len(lines)} lignes brutes -> {len(non_empty_lines)} lignes après nettoyage")
            
            # Détecter le séparateur automatiquement sur le contenu nettoyé
            sample_content = '\n'.join(non_empty_lines[:3])
            sniffer = csv.Sniffer()
            try:
                dialect = sniffer.sniff(sample_content, delimiters=';,\t|')
                detected_separator = dialect.delimiter
                print(f"  >>> Séparateur détecté pour {file_name}: '{detected_separator}'")
            except:
                detected_separator = ';'  # Fallback
                print(f"  >>> Échec détection séparateur pour {file_name}, utilisation de ';' par défaut")
            
            # Lire le CSV nettoyé
            if i == 0:
                df = pd.read_csv(io.StringIO(cleaned_content), sep=detected_separator, header=0, engine='python')
                column_names_from_first_csv = df.columns.tolist() # Sauvegarder les noms de colonnes
                print(f"  >>> Premier CSV ({file_name}): Noms de colonnes dérivés: {column_names_from_first_csv}")
                print(f"  >>> Premier CSV ({file_name}): Shape: {df.shape}")
                print(f"  >>> Premier CSV ({file_name}): Head(2):\\n{df.head(2).to_string()}")
            else:
                df = pd.read_csv(io.StringIO(cleaned_content), sep=detected_separator, header=None, names=column_names_from_first_csv, skiprows=1, engine='python')
                print(f"  >>> CSV Suivant ({file_name}): Shape après lecture (skiprows=1, names={column_names_from_first_csv}): {df.shape}")
                print(f"  >>> CSV Suivant ({file_name}): Head(2):\\n{df.head(2).to_string()}")
            
            all_data_frames.append(df)
        except Exception as e:
            print(f"Erreur lors de la lecture des données CSV du fichier {file_name} après décodage: {e}")
            # Selon la robustesse désirée, on pourrait sauter le fichier ou lever une erreur.
            continue
    
    if not all_data_frames:
        print("Aucune donnée CSV n'a été chargée.")
        return None, "error_no_data.xlsx" # Retourner None pour indiquer une erreur

    combined_df = pd.concat(all_data_frames, ignore_index=True)
    print(f"CSVs combinés. Nombre total de lignes après concaténation: {len(combined_df)}") # DEBUG: Nombre de lignes
    print(f"Dimensions du DataFrame après concaténation: {combined_df.shape}")
    print(f"Noms des colonnes du DataFrame combiné initial: {combined_df.columns.tolist()}")

    # --- AJOUT : Création de la feuille "Raw_Concatenated_Data" ---
    # Copier le DataFrame brut immédiatement après la concaténation
    df_raw_concat_info = combined_df.copy()

    # Définir les indices des colonnes basé sur les noms (plus robuste)
    # Ces noms doivent correspondre aux en-têtes réels de vos fichiers CSV.
    # Exemple: si vos colonnes sont nommées 'Diametre', 'CoeffG', 'CoeffH', 'CoeffI'
    # Adaptez ces noms à votre structure CSV exacte.
    # Pour l'instant, je suppose que les noms de colonnes du premier CSV sont représentatifs.
    # Si les noms de colonnes ne sont pas fiables/constants, il faudra utiliser des indices (iloc).
    
    # Tentative de mapper les colonnes VBA (E, G, H, I) à des noms ou indices.
    # VBA Col E (HA Type) -> Supposons que ce soit la 5ème colonne (index 4)
    # VBA Col G (pour calc K) -> Supposons 7ème col (index 6)
    # VBA Col H (pour calc K) -> Supposons 8ème col (index 7)
    # VBA Col I (pour calc K) -> Supposons 9ème col (index 8)
    try:
        col_E_idx, col_G_idx, col_H_idx, col_I_idx = 4, 6, 7, 8 # Indices supposés
        col_E_name = combined_df.columns[col_E_idx]
        col_G_name = combined_df.columns[col_G_idx]
        col_H_name = combined_df.columns[col_H_idx]
        col_I_name = combined_df.columns[col_I_idx]
        print(f"Mapping des colonnes utilisé : E (Type HA) -> '{col_E_name}', G -> '{col_G_name}', H -> '{col_H_name}', I -> '{col_I_name}'") # DEBUG
    except IndexError:
        print("Erreur: Le CSV ne contient pas assez de colonnes pour mapper E, G, H, I par indice.")
        return None, "error_column_mapping.xlsx"

    # 2. Nettoyer et Convertir les Données
    cols_to_clean_numeric = [col_G_name, col_H_name, col_I_name]
    for col_name in cols_to_clean_numeric:
        combined_df[col_name] = combined_df[col_name].apply(clean_value_for_numeric_conversion)
        combined_df[col_name] = pd.to_numeric(combined_df[col_name], errors='coerce')

    combined_df[col_E_name] = pd.to_numeric(combined_df[col_E_name], errors='coerce')
    
    # *** AUCUNE LIGNE NE DOIT ÊTRE SUPPRIMÉE ***
    # Au lieu de supprimer les lignes avec des NaN, on va les gérer intelligemment
    
    print(f"Avant traitement des valeurs manquantes: {len(combined_df)} lignes")
    
    # Traitement spécial pour chaque colonne critique :
    
    # Colonne E (Type HA) : Remplacer NaN par 0 (sera ignoré dans les calculs)
    combined_df[col_E_name] = combined_df[col_E_name].fillna(0)
    
    # Colonne G : Remplacer NaN par 0 
    combined_df[col_G_name] = combined_df[col_G_name].fillna(0)
    
    # Colonne H : Remplacer NaN par 0
    combined_df[col_H_name] = combined_df[col_H_name].fillna(0)
    
    # Colonne I : NE PAS remplacer les NaN - on va les traiter dans le calcul de K
    # combined_df[col_I_name] = combined_df[col_I_name].fillna(1)  # Supprimé
    
    print(f"Après traitement des valeurs manquantes: {len(combined_df)} lignes (aucune suppression)")
    print(f"Valeurs manquantes par colonne critique:")
    print(f"  {col_E_name}: {combined_df[col_E_name].isna().sum()}")
    print(f"  {col_G_name}: {combined_df[col_G_name].isna().sum()}")
    print(f"  {col_H_name}: {combined_df[col_H_name].isna().sum()}")
    print(f"  {col_I_name}: {combined_df[col_I_name].isna().sum()}")

    # 3. Calculer la Colonne K avec logique conditionnelle
    # Si I est vide (NaN) : K = G * H
    # Si I contient une valeur ≠ 0 : K = G * (H * 2 + I * 2 + 0.05)
    def calculate_k_conditional(row):
        G_val = row[col_G_name]
        H_val = row[col_H_name] 
        I_val = row[col_I_name]
        
        # Si I est NaN (vide), utiliser K = G * H
        if pd.isna(I_val):
            result = G_val * H_val
            return result
        else:
            # Si I a une valeur, utiliser la formule complète K = G * (H * 2 + I * 2 + 0.05)
            result = G_val * (H_val * 2 + I_val * 2 + 0.05)
            return result
    
    combined_df['K_Calculated'] = combined_df.apply(calculate_k_conditional, axis=1)
    print("Colonne 'K_Calculated' calculée avec logique conditionnelle :")
    print("  - Si I vide : K = G * H")
    print("  - Si I non vide : K = G * (H * 2 + I * 2 + 0.05)")

    # --- AJOUT : Création de la feuille "Info_Brute_K" ---
    # Copier le DataFrame avant le groupby pour la feuille d'info
    df_info_brute_k = combined_df.copy() # Renommé pour clarté

    # 4. Agréger par type d'armature (Colonne E)
    summary_by_ha = combined_df.groupby(col_E_name)['K_Calculated'].sum().reindex(HA_TYPES, fill_value=0)
    print("Synthèse par type d'HA (longueur développée):")
    print(summary_by_ha)

    # 5. Construire le DataFrame de résultat final
    result_data_rows = {
        "Armature": [
            "longeur en ml", "Nbr barre", "Prix/Tonne",
            "Nbr barre/tonne", "Prix par barre", "Prix"
        ]
    }
    total_prix_global = 0

    for ha_type in HA_TYPES:
        col_header_name = f"HA{ha_type}"
        longeur_ml = summary_by_ha.get(ha_type, 0)
        
        # Nbr barre: VBA = Ceiling(Ceiling(sum_longueur, 1) / 12, 1)
        nbr_barre = math.ceil(math.ceil(longeur_ml) / 12) if longeur_ml > 0 else 0
        
        prix_tonne_val = PRIX_TONNE
        nbr_barre_tonne_val = NBR_BARRE_TONNE_MAP.get(ha_type, 0)
        
        prix_par_barre_val = (prix_tonne_val / nbr_barre_tonne_val) if nbr_barre_tonne_val > 0 else 0
        # Le VBA ne semble pas explicitement arrondir ici, mais Excel peut le faire à l'affichage
        # Pour correspondre à l'image, un arrondi entier semble appliqué
        prix_par_barre_val = round(prix_par_barre_val) 

        prix_final_ha = nbr_barre * prix_par_barre_val
        total_prix_global += prix_final_ha
        
        result_data_rows[col_header_name] = [
            round(longeur_ml, 2),    # longueur en ml
            nbr_barre,               # Nbr barre
            prix_tonne_val,          # Prix/Tonne
            nbr_barre_tonne_val,     # Nbr barre/tonne
            prix_par_barre_val,      # Prix par barre
            round(prix_final_ha)     # Prix
        ]

    result_df = pd.DataFrame(result_data_rows)
    
    # 6. Générer le fichier Excel en mémoire
    output_excel_filename = "Synthese_Armatures.xlsx"
    wb = openpyxl.Workbook()
    ws_resultat = wb.active
    ws_resultat.title = "Resultat"

    # --- AJOUT : Écrire la feuille "Raw_Concatenated_Data" ---
    ws_raw_concat = wb.create_sheet(title="Raw_Concatenated_Data")
    # Écrire les en-têtes pour la feuille de données brutes concaténées
    if not df_raw_concat_info.empty:
        ws_raw_concat.append(df_raw_concat_info.columns.tolist())
        # Écrire les données
        for r_idx, row in enumerate(df_raw_concat_info.itertuples(index=False), 1):
            for c_idx, value in enumerate(row, 1):
                ws_raw_concat.cell(row=r_idx + 1, column=c_idx, value=value)
        print(f"Feuille 'Raw_Concatenated_Data' ajoutée avec {len(df_raw_concat_info)} lignes.")
    else:
        print("Le DataFrame df_raw_concat_info est vide, la feuille 'Raw_Concatenated_Data' ne sera pas remplie.")

    # --- AJOUT : Écrire la feuille "Info_Brute_K" --- 
    ws_info_k = wb.create_sheet(title="Info_Brute_K") # Nom de variable ws_info_k pour clarté
    # Écrire les en-têtes pour la feuille info_k
    if not df_info_brute_k.empty:
        ws_info_k.append(df_info_brute_k.columns.tolist())
        # Écrire les données
        for r_idx, row in enumerate(df_info_brute_k.itertuples(index=False), 1):
            for c_idx, value in enumerate(row, 1):
                ws_info_k.cell(row=r_idx + 1, column=c_idx, value=value)
        print(f"Feuille 'Info_Brute_K' ajoutée avec {len(df_info_brute_k)} lignes.")
    else:
        print("Le DataFrame df_info_brute_k est vide, la feuille 'Info_Brute_K' ne sera pas remplie.")

    # Écrire les en-têtes du tableau de synthèse sur la feuille "Resultat"
    header_list = ["Armature"] + [f"HA{ha}" for ha in HA_TYPES] + ["Prix total du fer"]
    ws_resultat.append(header_list)

    # Écrire les données du DataFrame de synthèse
    for index in range(len(result_df["Armature"])):
        row_to_append = [result_df["Armature"][index]]
        for ha_type in HA_TYPES:
            col_header_name = f"HA{ha_type}"
            row_to_append.append(result_df[col_header_name][index])
        
        # Ajouter le prix total global sur la dernière ligne "Prix"
        if result_df["Armature"][index] == "Prix":
            row_to_append.append(round(total_prix_global))
        else:
            row_to_append.append(None) # Cellule vide pour les autres lignes dans cette colonne
        ws_resultat.append(row_to_append)

    # Appliquer le formatage (inspiré du VBA)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Bleu
    for cell in ws_resultat[1]: # Première ligne (en-têtes)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx in range(2, ws_resultat.max_row + 1):
        ws_resultat[f'A{row_idx}'].font = Font(bold=True) # Colonne "Armature" en gras

    # Format des nombres (approximatif du style "Millier" et spécifique)
    # Colonnes B à F (HA6 à HA14)
    for col_idx_letter in [get_column_letter(c) for c in range(2, 2 + len(HA_TYPES))]:
        for row_idx in range(2, ws_resultat.max_row + 1):
            cell = ws_resultat[f'{col_idx_letter}{row_idx}']
            if isinstance(cell.value, (int, float)):
                armature_label = ws_resultat[f'A{row_idx}'].value
                if armature_label == "longeur en ml":
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '#,##0' # Entier avec séparateur de milliers
    
    # Cellule du prix total global (G7 dans l'image, dernière ligne/colonne ici)
    cell_prix_total_global = ws_resultat.cell(row=ws_resultat.max_row, column=len(header_list))
    cell_prix_total_global.font = Font(bold=True, color="FFFFFF", size=14)
    cell_prix_total_global.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # Rouge
    cell_prix_total_global.number_format = '#,##0'

    # Ajuster largeur des colonnes
    ws_resultat.column_dimensions['A'].width = 20
    for i, _ in enumerate(HA_TYPES):
        ws_resultat.column_dimensions[get_column_letter(2 + i)].width = 15
    ws_resultat.column_dimensions[get_column_letter(len(header_list))].width = 18

    # Sauvegarder dans un flux binaire
    output_io = io.BytesIO()
    wb.save(output_io)
    output_io.seek(0)
    
    print(f"Fichier Excel '{output_excel_filename}' généré en mémoire.")
    return output_io, output_excel_filename


if __name__ == '__main__':
    # Bloc de test pour exécuter le script localement
    print("Exécution du script combineArm.py en mode test...")
    # Simuler des fichiers CSV venant de Flutter
    # IMPORTANT: Adaptez ces données et surtout les NOMS DE COLONNES aux vôtres!
    # Si vos CSV n'ont pas d'en-têtes, la lecture devra être ajustée (header=None, names=[...])
    dummy_csv_content1 = (
        "ColA;ColB;ColC;ColD;TypeHA;Esp;ValG;ValH;ValI;Autre\n"
        "id1;b;c;d;6;15;A = 10.5;B = 0,25;C = 0,35;xxx\n"
        "id2;b;c;d;8;20;20;0.30;0.40;yyy\n"
        "id3;b;c;d;6;10;C = 15,5;0,18;0,28;zzz"
    )
    dummy_csv_content2 = (
        "ColA;ColB;ColC;ColD;TypeHA;Esp;ValG;ValH;ValI;Autre\n"
        "id4;b;c;d;10;15;30;0.3;0.4;aaa\n"
        "id5;b;c;d;12;12;B = 25.0; A = 0,22;C = 0,32;bbb\n"
        "id6;b;c;d;6;10;12;0.2;0.3;ccc"
    )
    
    files_from_flutter = [
        {'name': 'fichier1.csv', 'bytes': dummy_csv_content1.encode('utf-8')},
        {'name': 'fichier2.csv', 'bytes': dummy_csv_content2.encode('utf-8')}
    ]
    
    result_io, result_filename = process_armature_csvs(files_from_flutter)
    
    if result_io:
        # Sauvegarder localement pour inspection
        with open(result_filename, 'wb') as f_out:
            f_out.write(result_io.getvalue())
        print(f"Fichier de test sauvegardé: {result_filename}")
    else:
        print("Échec de la génération du fichier de test.") 